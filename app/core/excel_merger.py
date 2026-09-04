from __future__ import annotations

import logging
import re
import time
from copy import copy
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from app.config import EXCEL_MAX_SHEET_NAME
from app.core.models import MergeOptions, MergeResult, SheetItem

INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
ProgressCallback = Callable[[int, int, str], None]


class MergeError(Exception):
    pass


class MergeCancelled(MergeError):
    pass


def validate_sheet_name(name: str) -> str | None:
    if not name.strip():
        return "Tên sheet không được để trống."
    if len(name) > EXCEL_MAX_SHEET_NAME:
        return "Tên sheet không được dài quá 31 ký tự."
    if INVALID_SHEET_CHARS.search(name):
        return "Tên sheet không được chứa : \\ / ? * [ ]"
    if name.startswith("'") or name.endswith("'"):
        return "Tên sheet không được bắt đầu hoặc kết thúc bằng dấu nháy đơn."
    return None


def unique_sheet_name(name: str, used: set[str]) -> str:
    base = INVALID_SHEET_CHARS.sub("_", name).strip("'")[:EXCEL_MAX_SHEET_NAME] or "Sheet"
    lowered = {item.casefold() for item in used}
    if base.casefold() not in lowered:
        return base
    index = 2
    while True:
        suffix = f"_{index}"
        candidate = f"{base[:EXCEL_MAX_SHEET_NAME - len(suffix)]}{suffix}"
        if candidate.casefold() not in lowered:
            return candidate
        index += 1


class ExcelMerger:
    def merge(self, items: list[SheetItem], output_path: Path, options: MergeOptions,
              progress: ProgressCallback | None = None,
              cancelled: Callable[[], bool] | None = None) -> MergeResult:
        started = time.monotonic()
        selected = [item for item in items if item.selected]
        if not selected:
            raise MergeError("Chưa chọn sheet nào để ghép.")
        output_path = Path(output_path)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        used: set[str] = set()
        resolved: list[tuple[SheetItem, str]] = []
        for item in selected:
            error = validate_sheet_name(item.output_name)
            if error and not options.auto_resolve_duplicates:
                raise MergeError(f"{item.output_name}: {error}")
            name = unique_sheet_name(item.output_name, used) if options.auto_resolve_duplicates else item.output_name
            if name.casefold() in {value.casefold() for value in used}:
                raise MergeError(f"Tên sheet bị trùng: {name}")
            used.add(name)
            resolved.append((item, name))

        target = Workbook()
        target.remove(target.active)
        source_books: dict[tuple[Path, bool], object] = {}
        warnings: list[str] = []
        copied_count = 0
        try:
            for index, (item, output_name) in enumerate(resolved, 1):
                if cancelled and cancelled():
                    raise MergeCancelled("Đã hủy quá trình ghép.")
                key = (item.source_path, options.values_only)
                if key not in source_books:
                    try:
                        source_books[key] = load_workbook(
                            item.source_path, data_only=options.values_only,
                            keep_vba=item.source_path.suffix.lower() == ".xlsm", keep_links=True)
                    except PermissionError as exc:
                        raise MergeError(f"Không thể đọc {item.source_path.name}. Hãy đóng file trong Excel.") from exc
                    except Exception as exc:
                        raise MergeError(f"Không thể đọc {item.source_path.name}: {exc}") from exc
                source = source_books[key]
                if item.source_sheet not in source.sheetnames:
                    raise MergeError(f"Không tìm thấy sheet {item.source_sheet} trong {item.source_path.name}.")
                sheet = source[item.source_sheet]
                if options.skip_empty_sheets and self._is_empty(sheet):
                    warnings.append(f"Đã bỏ qua sheet trống: {item.source_sheet}")
                    continue
                destination = target.create_sheet(output_name)
                self._copy_sheet(sheet, destination, options)
                copied_count += 1
                if progress:
                    progress(index, len(resolved), item.source_path.name)
            if not target.worksheets:
                raise MergeError("Không có sheet có dữ liệu để ghi vào file kết quả.")
            # Excel yêu cầu ít nhất một worksheet hiển thị.
            if all(sheet.sheet_state != "visible" for sheet in target.worksheets):
                target.worksheets[0].sheet_state = "visible"
                warnings.append("Sheet đầu tiên được đặt hiển thị vì Excel yêu cầu ít nhất một sheet hiển thị.")
            try:
                target.save(output_path)
            except PermissionError as exc:
                raise MergeError("Không thể ghi file kết quả. Hãy đóng file nếu đang mở trong Excel.") from exc
        finally:
            for workbook in source_books.values():
                workbook.close()
            target.close()
        return MergeResult(output_path, len({item.source_path for item in selected}),
                           copied_count, time.monotonic() - started, warnings)

    @staticmethod
    def _is_empty(sheet) -> bool:
        return not any(cell.value is not None for row in sheet.iter_rows() for cell in row)

    def _copy_sheet(self, source, target, options: MergeOptions) -> None:
        leading = 0
        if options.trim_leading_empty_rows:
            for row in source.iter_rows():
                if any(cell.value is not None for cell in row):
                    break
                leading += 1
        for row in source.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell) or cell.row <= leading:
                    continue
                new = target.cell(cell.row - leading, cell.column, cell.value)
                if options.preserve_formatting and cell.has_style:
                    new.font = copy(cell.font)
                    new.fill = copy(cell.fill)
                    new.border = copy(cell.border)
                    new.alignment = copy(cell.alignment)
                    new.protection = copy(cell.protection)
                    new.number_format = cell.number_format
                if cell.hyperlink:
                    new._hyperlink = copy(cell.hyperlink)
                if cell.comment:
                    new.comment = copy(cell.comment)
        if options.preserve_formatting:
            for key, dimension in source.column_dimensions.items():
                target.column_dimensions[key] = copy(dimension)
            for index, dimension in source.row_dimensions.items():
                if index > leading:
                    copied = copy(dimension)
                    copied.index = index - leading
                    target.row_dimensions[index - leading] = copied
            for merged in source.merged_cells.ranges:
                if merged.min_row > leading:
                    target.merge_cells(start_row=merged.min_row - leading, start_column=merged.min_col,
                                       end_row=merged.max_row - leading, end_column=merged.max_col)
            for image in getattr(source, "_images", []):
                try:
                    target.add_image(copy(image), copy(image.anchor))
                except Exception:
                    logging.getLogger(__name__).warning("Could not copy an image from %s", source.title)
        target.sheet_format = copy(source.sheet_format)
        target.sheet_properties = copy(source.sheet_properties)
        target.page_margins = copy(source.page_margins)
        target.page_setup = copy(source.page_setup)
        target.print_options = copy(source.print_options)
        target.views = copy(source.views)
        target.freeze_panes = source.freeze_panes
        target.auto_filter = copy(source.auto_filter)
        target.sheet_state = source.sheet_state
        target.print_area = source.print_area
        target.print_title_cols = source.print_title_cols
        target.print_title_rows = source.print_title_rows
