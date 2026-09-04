from pathlib import Path

from copy import copy

from openpyxl import Workbook, load_workbook

from app.core.excel_merger import ExcelMerger, unique_sheet_name, validate_sheet_name
from app.core.models import MergeOptions, SheetItem


def make_book(path: Path, title: str, value: str) -> None:
    book = Workbook(); sheet = book.active; sheet.title = title
    sheet["A1"] = value; sheet["B2"] = "=1+2"
    font = copy(sheet["A1"].font); font.bold = True; sheet["A1"].font = font
    sheet.column_dimensions["A"].width = 24; sheet.freeze_panes = "A2"; sheet.merge_cells("C3:D3")
    book.save(path); book.close()


def test_merge_keeps_content_and_order(tmp_path: Path) -> None:
    first, second, output = tmp_path / "a.xlsx", tmp_path / "b.xlsx", tmp_path / "out.xlsx"
    make_book(first, "Data", "A"); make_book(second, "Data", "B")
    items = [SheetItem(first, "Data", "Data"), SheetItem(second, "Data", "Data")]
    result = ExcelMerger().merge(items, output, MergeOptions())
    book = load_workbook(output, data_only=False)
    assert book.sheetnames == ["Data", "Data_2"]
    assert book["Data"]["A1"].value == "A" and book["Data"]["B2"].value == "=1+2"
    assert book["Data"]["A1"].font.bold and book["Data"].column_dimensions["A"].width == 24
    assert book["Data"].freeze_panes == "A2" and "C3:D3" in book["Data"].merged_cells
    assert result.sheet_count == 2
    book.close()


def test_name_validation_and_uniqueness() -> None:
    assert validate_sheet_name("bad/name") is not None
    assert validate_sheet_name("Valid") is None
    assert unique_sheet_name("Data", {"data"}) == "Data_2"
